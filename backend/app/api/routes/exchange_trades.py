"""
Endpoints para sincronizar y gestionar trades importados desde el exchange.
"""
import uuid
from datetime import datetime, timezone
from decimal import Decimal

from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.dependencies import get_current_user_id
from app.exchanges.factory import create_exchange
from app.models.exchange_account import ExchangeAccount
from app.models.exchange_trade import ExchangeTrade
from app.models.position import Position
from app.schemas.exchange_trade import (
    ExchangeTradeFilter,
    ExchangeTradeResponse,
    ExchangeTradeSyncResult,
)
from app.services.database import get_db

router = APIRouter(prefix="/exchange-trades", tags=["exchange-trades"])


@router.post("/sync/{account_id}", response_model=ExchangeTradeSyncResult)
async def sync_trades(
    account_id: uuid.UUID,
    days: int = Query(30, ge=1, le=365),
    user_id: uuid.UUID = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db),
):
    """
    Sincroniza trades históricos desde el exchange a la base de datos.
    
    - Para trades que coinciden con posiciones del bot: source='bot'
    - Para trades sin coincidencia: source='manual'
    """
    from loguru import logger
    
    # Verificar acceso a la cuenta
    result = await db.execute(
        select(ExchangeAccount).where(
            ExchangeAccount.id == account_id,
            ExchangeAccount.user_id == user_id,
        )
    )
    account = result.scalar_one_or_none()
    if not account:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Cuenta no encontrada")
    
    # Crear cliente del exchange
    try:
        exchange = create_exchange(account)
    except Exception as e:
        logger.error(f"Error creando exchange client: {e}")
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST, 
            f"Error conectando al exchange: {str(e)}"
        )
    
    try:
        # Calcular timestamp desde (ms)
        since_ts = int((datetime.now(timezone.utc).timestamp() - days * 24 * 3600) * 1000)
        since_dt = datetime.fromtimestamp(since_ts / 1000, tz=timezone.utc)
        logger.info(f"Sync trades: account={account.exchange}, since={since_dt}, days={days}")
        
        # Obtener IDs ya existentes en BD para evitar duplicados (NO borrar nada)
        existing_result = await db.execute(
            select(ExchangeTrade.exchange_trade_id).where(
                ExchangeTrade.exchange_account_id == account_id,
                ExchangeTrade.user_id == user_id,
            )
        )
        existing_ids = {row[0] for row in existing_result.all()}
        logger.info(f"Sync trades: {len(existing_ids)} trades ya en BD")

        # Obtener trades del exchange
        trades_raw = await exchange.get_trade_history(limit=500, since=since_ts)
        logger.info(f"Sync trades: obtenidos {len(trades_raw)} trades del exchange")
        
        if not trades_raw:
            return ExchangeTradeSyncResult(total_synced=0, new_trades=0, updated_trades=0)
        
        # Posiciones del usuario (abiertas Y cerradas) para clasificar source=bot
        from app.models.bot_config import BotConfig
        positions_result = await db.execute(
            select(Position)
            .join(BotConfig, Position.bot_id == BotConfig.id)
            .where(BotConfig.user_id == user_id)
        )
        positions = positions_result.scalars().all()

        # Mapear posiciones por símbolo para lookup rápido
        positions_by_symbol: dict[str, list] = {}
        for pos in positions:
            positions_by_symbol.setdefault(pos.symbol, []).append(pos)
        
        new_count = 0
        errors = []
        
        for trade_data in trades_raw:
            try:
                trade_id = str(trade_data.get("id", ""))
                if not trade_id:
                    continue
                
                symbol = trade_data.get("symbol", "")
                trade_ts = trade_data.get("timestamp")
                
                # Verificar que el trade esté dentro del período solicitado
                if trade_ts and trade_ts < since_ts:
                    continue

                # Saltar si ya existe en BD
                if trade_id in existing_ids:
                    continue

                # Determinar source y relaciones
                source = "manual"
                position_id = None
                bot_id = None
                
                # Buscar posición coincidente (por símbolo y ventana temporal)
                for pos in positions_by_symbol.get(symbol, []):
                    if trade_ts and pos.opened_at:
                        trade_dt = datetime.fromtimestamp(trade_ts / 1000, tz=timezone.utc)
                        end = pos.closed_at or datetime.now(timezone.utc)
                        if pos.opened_at <= trade_dt <= end:
                            source = "bot"
                            position_id = pos.id
                            bot_id = pos.bot_id
                            break
                
                # Crear nuevo trade
                closed_at = datetime.fromtimestamp(
                    trade_ts / 1000, tz=timezone.utc
                ) if trade_ts else datetime.now(timezone.utc)
                
                new_trade = ExchangeTrade(
                    user_id=user_id,
                    exchange_account_id=account_id,
                    position_id=position_id,
                    bot_id=bot_id,
                    source=source,
                    exchange_trade_id=trade_id,
                    symbol=symbol,
                    side=trade_data.get("side") or "long",
                    quantity=Decimal(str(trade_data.get("quantity") or 0)),
                    entry_price=trade_data.get("price"),
                    exit_price=trade_data.get("price"),
                    realized_pnl=trade_data.get("pnl"),  # Puede ser None
                    fee=Decimal(str(trade_data.get("fee") or 0)),
                    fee_asset=trade_data.get("fee_asset") or "USDT",
                    closed_at=closed_at,
                    order_type=trade_data.get("order_type") or "market",
                    status="closed",
                    raw_data=str(trade_data.get("raw", {}))[:1000],  # Limitar tamaño
                )
                db.add(new_trade)
                new_count += 1
                    
            except Exception as e:
                error_msg = f"Error procesando trade {trade_data.get('id')}: {str(e)}"
                logger.error(error_msg)
                errors.append(error_msg)
        
        await db.commit()
        
        logger.info(f"Sync trades: {new_count} nuevos trades guardados, {len(errors)} errores")

        return ExchangeTradeSyncResult(
            total_synced=len(trades_raw),
            new_trades=new_count,
            updated_trades=0,
            errors=errors if errors else None,
        )
        
    except Exception as e:
        raise HTTPException(
            status.HTTP_502_BAD_GATEWAY,
            f"Error sincronizando trades: {str(e)}"
        )
    finally:
        await exchange.close()


@router.get("", response_model=list[ExchangeTradeResponse])
async def list_trades(
    account_id: uuid.UUID | None = Query(None),
    bot_id: uuid.UUID | None = Query(None),
    source: str | None = Query(None, regex="^(bot|manual)$"),
    symbol: str | None = Query(None),
    days: int = Query(30, ge=1, le=365),
    limit: int = Query(100, ge=1, le=500),
    user_id: uuid.UUID = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db),
):
    """
    Lista trades importados del exchange con filtros.
    
    Filters:
    - source='bot': Solo trades ejecutados por el bot
    - source='manual': Solo trades manuales del usuario
    - source=None: Todos los trades
    """
    # Calcular fecha de inicio como días calendario (consistente con frontend)
    # days=7 significa: desde hace 6 días hasta hoy (7 días total)
    from datetime import timedelta
    today = datetime.now(timezone.utc).date()
    start_date = today - timedelta(days=days - 1)
    since_dt = datetime.combine(start_date, datetime.min.time(), tzinfo=timezone.utc)
    
    query = select(ExchangeTrade).where(
        ExchangeTrade.user_id == user_id,
        ExchangeTrade.closed_at >= since_dt,
    )
    
    if account_id:
        query = query.where(ExchangeTrade.exchange_account_id == account_id)
    if bot_id:
        query = query.where(ExchangeTrade.bot_id == bot_id)
    if source:
        query = query.where(ExchangeTrade.source == source)
    if symbol:
        query = query.where(ExchangeTrade.symbol.ilike(f"%{symbol}%"))
    
    query = query.order_by(ExchangeTrade.closed_at.desc()).limit(limit)
    
    result = await db.execute(query)
    return result.scalars().all()


@router.get("/stats", response_model=dict)
async def get_trade_stats(
    account_id: uuid.UUID | None = Query(None),
    days: int = Query(30, ge=1, le=365),
    user_id: uuid.UUID = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db),
):
    """
    Estadísticas de trades separadas por source (bot vs manual).
    """
    from sqlalchemy import func
    from datetime import timedelta
    
    # Calcular fecha de inicio como días calendario (consistente con frontend)
    today = datetime.now(timezone.utc).date()
    start_date = today - timedelta(days=days - 1)
    since_dt = datetime.combine(start_date, datetime.min.time(), tzinfo=timezone.utc)
    
    # Base query
    base_query = select(ExchangeTrade).where(
        ExchangeTrade.user_id == user_id,
        ExchangeTrade.closed_at >= since_dt,
    )
    
    if account_id:
        base_query = base_query.where(ExchangeTrade.exchange_account_id == account_id)
    
    # Stats por source
    stats = {
        "bot": {"count": 0, "total_pnl": 0.0, "winners": 0, "losers": 0},
        "manual": {"count": 0, "total_pnl": 0.0, "winners": 0, "losers": 0},
        "total": {"count": 0, "total_pnl": 0.0, "winners": 0, "losers": 0},
    }
    
    for source in ["bot", "manual"]:
        source_query = base_query.where(ExchangeTrade.source == source)
        
        result = await db.execute(source_query)
        trades = result.scalars().all()
        
        count = len(trades)
        total_pnl = sum(float(t.realized_pnl or 0) for t in trades)
        winners = sum(1 for t in trades if (t.realized_pnl or 0) > 0)
        losers = sum(1 for t in trades if (t.realized_pnl or 0) < 0)
        
        stats[source] = {
            "count": count,
            "total_pnl": round(total_pnl, 2),
            "winners": winners,
            "losers": losers,
            "win_rate": round(winners / count * 100, 1) if count > 0 else 0,
        }
        
        # Acumular en total
        stats["total"]["count"] += count
        stats["total"]["total_pnl"] += total_pnl
        stats["total"]["winners"] += winners
        stats["total"]["losers"] += losers
    
    # Calcular win_rate total
    if stats["total"]["count"] > 0:
        stats["total"]["win_rate"] = round(
            stats["total"]["winners"] / stats["total"]["count"] * 100, 1
        )
    
    stats["period_days"] = days
    stats["account_id"] = str(account_id) if account_id else None
    
    return stats


@router.post("/import-csv/{account_id}")
async def import_csv_trades(
    account_id: uuid.UUID,
    file: UploadFile = File(...),
    user_id: uuid.UUID = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db),
):
    """
    Importa trades desde un archivo exportado de BingX.
    Acepta XLSX (Excel) y CSV. BingX exporta XLSX con extensión .csv.
    """
    import csv
    import io
    from loguru import logger

    # Verificar acceso a la cuenta
    result = await db.execute(
        select(ExchangeAccount).where(
            ExchangeAccount.id == account_id,
            ExchangeAccount.user_id == user_id,
        )
    )
    account = result.scalar_one_or_none()
    if not account:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Cuenta no encontrada")

    # Cargar posiciones del usuario para clasificación
    from app.models.bot_config import BotConfig
    positions_result = await db.execute(
        select(Position).join(BotConfig).where(BotConfig.user_id == user_id)
    )
    positions = positions_result.scalars().all()

    positions_by_symbol: dict = {}
    for pos in positions:
        key = (pos.symbol, account.exchange)
        positions_by_symbol.setdefault(key, []).append(pos)

    # Leer bytes del archivo subido
    raw_bytes = await file.read()
    logger.info(f"Importando archivo '{file.filename}' ({len(raw_bytes)} bytes) para cuenta {account_id}")

    # Detectar formato: XLSX empieza con PK (ZIP magic bytes)
    is_xlsx = raw_bytes[:2] == b'PK'

    rows: list[dict] = []

    if is_xlsx:
        try:
            import openpyxl
            # Sin read_only para mejor compatibilidad con exportaciones de BingX
            wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)

            # Buscar la hoja con más datos
            ws = wb.active
            for sheet in wb.worksheets:
                if sheet.max_row > (ws.max_row or 0):
                    ws = sheet

            # Leer todas las filas como listas de valores
            all_rows = []
            for excel_row in ws.iter_rows(values_only=True):
                # Convertir cada celda a string, ignorar filas completamente vacías
                cells = [str(c).strip() if c is not None else "" for c in excel_row]
                if any(cells):  # al menos una celda no vacía
                    all_rows.append(cells)

            logger.info(f"XLSX: {len(all_rows)} filas brutas. Primera fila: {all_rows[0] if all_rows else 'vacío'}")

            if not all_rows:
                raise HTTPException(status.HTTP_400_BAD_REQUEST, "El archivo XLSX está vacío")

            # Encontrar la fila de cabeceras (la primera que tenga más de 1 columna con datos)
            header_idx = 0
            for i, row in enumerate(all_rows):
                non_empty = sum(1 for c in row if c)
                if non_empty > 1:
                    header_idx = i
                    break

            headers = all_rows[header_idx]
            logger.info(f"XLSX detectado. Headers (fila {header_idx}): {headers}. Filas de datos: {len(all_rows) - header_idx - 1}")

            for data_row in all_rows[header_idx + 1:]:
                if not any(data_row):
                    continue
                row_dict = {headers[i]: data_row[i] if i < len(data_row) else "" for i in range(len(headers))}
                rows.append(row_dict)

            wb.close()
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Error leyendo XLSX: {e}")
    else:
        try:
            text = raw_bytes.decode("utf-8-sig", errors="replace").replace('\r\n', '\n').replace('\r', '\n')
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
            logger.info(f"CSV detectado. Headers: {reader.fieldnames}. Filas: {len(rows)}")
        except Exception as e:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Error leyendo CSV: {e}")

    if not rows:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "El archivo está vacío o no tiene datos válidos")

    imported = 0
    skipped = 0
    errors: list[str] = []

    for row in rows:
        try:
            # ── Extraer campos (BingX Transaction History / Income History) ──
            # Columnas comunes según exportación BingX:
            # "Time(UTC type)", "Details", "Amount", "newAvailaAssets", "Futures"
            # O en Order History: "Time", "Symbol", "Side", "Price", "Qty", "Realized Profit", "Fee"

            def get_col(*keys):
                for k in keys:
                    # Buscar también con coincidencia parcial (ignora UTC+8 vs UTC type)
                    v = row.get(k)
                    if v is None:
                        for rk in row:
                            if rk.lower().startswith(k.lower()[:8]):
                                v = row[rk]
                                break
                    if v is not None and str(v).strip() not in ('', 'None', 'nan'):
                        return str(v).strip()
                return ''

            time_str = get_col('Time(UTC+8)', 'Time(UTC type)', 'Time', 'time', 'Date', 'date')
            details  = get_col('Details', 'details', 'Type', 'type')
            amount_str = get_col('Amount', 'amount', 'Realized Profit', 'realized_profit', 'PnL', 'pnl')
            symbol   = get_col('Futures', 'futures', 'Symbol', 'symbol', 'Contract', 'contract')
            side_str = get_col('Side', 'side', 'Direction', 'direction')
            qty_str  = get_col('Qty', 'qty', 'Quantity', 'quantity', 'Size', 'size', 'Filled', 'filled')
            price_str = get_col('Price', 'price', 'Avg Price', 'avg_price')
            fee_str   = get_col('Fee', 'fee', 'Trading Fee', 'trading_fee', 'Commission', 'commission')

            # Saltar filas sin símbolo
            if not symbol:
                skipped += 1
                continue

            # PnL
            try:
                profit = Decimal(str(amount_str).replace(',', ''))
            except Exception:
                profit = Decimal("0")

            # Fee
            try:
                fee = abs(Decimal(str(fee_str).replace(',', '')))
            except Exception:
                fee = Decimal("0")

            # Si Details dice que es fee y no hay columna Fee separada, tratar Amount como fee
            if fee == Decimal("0") and details and ('fee' in details.lower() or 'funding' in details.lower()):
                fee = abs(profit)
                profit = Decimal("0")

            # Parsear fecha
            closed_at = datetime.now(timezone.utc)
            if time_str:
                for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d', '%Y/%m/%d %H:%M:%S', '%Y/%m/%d'):
                    try:
                        closed_at = datetime.strptime(time_str[:19], fmt).replace(tzinfo=timezone.utc)
                        break
                    except ValueError:
                        continue

            # Normalizar símbolo: ARB-USDT → ARB/USDT:USDT
            symbol = symbol.replace(' ', '').upper()
            if '-' in symbol and '/' not in symbol:
                symbol = symbol.replace('-', '/')
            if '/' in symbol and ':' not in symbol and 'USDT' in symbol:
                symbol = symbol + ':USDT'
            elif '/' not in symbol and ':' not in symbol and symbol.endswith('USDT'):
                base = symbol[:-4]
                symbol = f"{base}/USDT:USDT"

            # Side
            side = "long"
            if side_str:
                sl = side_str.lower()
                if any(w in sl for w in ('sell', 'short', 'close long', 'venta')):
                    side = "short"

            # Cantidad y precio
            try:
                quantity = Decimal(str(qty_str).replace(',', '')) if qty_str else Decimal("0")
            except Exception:
                quantity = Decimal("0")
            try:
                price = Decimal(str(price_str).replace(',', '')) if price_str else None
            except Exception:
                price = None

            # ID único para deduplicación
            trade_id = f"{symbol.replace('/', '-').replace(':', '-')}_{closed_at.strftime('%Y%m%d%H%M%S')}_{str(profit).replace('.', '_').replace('-', 'n')}"

            # Verificar duplicado
            existing = await db.execute(
                select(ExchangeTrade).where(
                    ExchangeTrade.exchange_account_id == account_id,
                    ExchangeTrade.exchange_trade_id == trade_id,
                )
            )
            if existing.scalar_one_or_none():
                skipped += 1
                continue

            # Clasificar source (bot vs manual)
            source = "manual"
            position_id = None
            bot_id = None
            pos_key = (symbol, account.exchange)
            for pos in positions_by_symbol.get(pos_key, []):
                if pos.opened_at and pos.opened_at <= closed_at:
                    if pos.closed_at is None or closed_at <= pos.closed_at:
                        source = "bot"
                        position_id = pos.id
                        bot_id = pos.bot_id
                        break

            trade = ExchangeTrade(
                id=uuid.uuid4(),
                user_id=user_id,
                exchange_account_id=account_id,
                position_id=position_id,
                bot_id=bot_id,
                source=source,
                exchange_trade_id=trade_id,
                symbol=symbol,
                side=side,
                quantity=quantity,
                entry_price=price,
                exit_price=price,
                realized_pnl=profit,
                fee=fee,
                closed_at=closed_at,
                status="closed",
            )
            db.add(trade)
            imported += 1

        except Exception as e:
            errors.append(f"Fila omitida: {e}")
            continue

    await db.commit()
    logger.info(f"Importación completada: {imported} importados, {skipped} omitidos, {len(errors)} errores")

    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors[:10],
        "account": account.label,
        "format": "xlsx" if is_xlsx else "csv",
    }
